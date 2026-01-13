#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分析AI回答为空的原因
检查代码逻辑和可能的问题点
"""
import pandas as pd
import openpyxl

def analyze_code_issue():
    """分析代码中可能导致空值的问题"""
    print('='*80)
    print('代码逻辑分析：可能导致AI回答为空的原因')
    print('='*80)
    
    issues = []
    
    print('\n1. DeepSeek API analyze_case方法（utils/deepseek_api.py）:')
    print('   - 第272行: answer = message.get(\'content\', \'\')')
    print('   - 如果API返回的content字段为空，answer会是空字符串')
    print('   - 第276-282行: 如果content为空，代码会警告但不处理，保持answer为空')
    print('   ⚠️  问题：API可能返回content为空的情况（如内容过滤、截断等）')
    issues.append('API返回content为空')
    
    print('\n2. process_cases.py process_single_question函数:')
    print('   - 第191行: ai_response = thread_ai_api.analyze_case(...)')
    print('   - 第196-201行: 提取answer')
    print('   - 第203行: result[\'AI回答\'] = ai_answer')
    print('   ⚠️  问题：如果ai_answer是空字符串，仍然会被设置到result中')
    issues.append('空字符串被设置到result')
    
    print('\n3. 异常处理（process_cases.py 第251-254行）:')
    print('   - 如果API调用抛出异常，会捕获并返回result')
    print('   - 但此时result[\'AI回答\']可能还未设置，或设置为空')
    print('   ⚠️  问题：异常时result[\'AI回答\']可能是None或未定义')
    issues.append('异常处理时未正确设置AI回答')
    
    print('\n4. _make_request方法返回None的情况（utils/deepseek_api.py 第198行）:')
    print('   - 如果所有重试都失败，_make_request返回None')
    print('   - analyze_case在第255行检查response，如果为None会抛出异常')
    print('   ⚠️  问题：如果异常被捕获，可能导致空值')
    issues.append('API请求失败返回None')
    
    print('\n5. 并发处理问题（process_cases.py 第260-264行）:')
    print('   - 使用线程池并发处理多个问题')
    print('   - 如果某个线程的异常未被正确处理，可能导致结果丢失')
    print('   ⚠️  问题：并发环境下异常处理可能不完整')
    issues.append('并发处理异常')
    
    return issues

def check_excel_file():
    """检查Excel文件中的空值情况"""
    file_path = 'data/results_20260112_105927/5个案例_统一评估结果_108cases.xlsx'
    
    print('\n' + '='*80)
    print('检查Excel文件中的空值')
    print('='*80)
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb['DeepSeek']
    
    # 检查J7单元格
    cell_j7 = ws['J7']
    print(f'\nJ7单元格:')
    print(f'  值: {repr(cell_j7.value)}')
    print(f'  是否为None: {cell_j7.value is None}')
    
    # 检查该案例的所有行
    case_id = 'case_20251230_101232_6'
    case_rows = []
    for row_idx in range(2, ws.max_row + 1):
        if ws[f'A{row_idx}'].value == case_id:
            case_rows.append(row_idx)
    
    print(f'\n案例 {case_id} 的所有行: {case_rows}')
    for row_num in case_rows:
        cell_j = ws[f'J{row_num}']
        question_num = ws[f'D{row_num}'].value
        print(f'  行{row_num} (问题{question_num}): AI回答={"空" if cell_j.value is None or str(cell_j.value).strip() == "" else "有内容"}')
    
    wb.close()

def suggest_fixes():
    """建议修复方案"""
    print('\n' + '='*80)
    print('建议的修复方案')
    print('='*80)
    
    print('\n1. 在DeepSeek API的analyze_case方法中:')
    print('   - 如果content为空，应该抛出异常而不是返回空字符串')
    print('   - 或者重试API调用')
    
    print('\n2. 在process_cases.py的process_single_question函数中:')
    print('   - 检查ai_answer是否为空，如果为空应该重试或记录错误')
    print('   - 异常处理时确保result[\'AI回答\']有默认值')
    
    print('\n3. 添加日志记录:')
    print('   - 记录所有API调用失败的情况')
    print('   - 记录content为空的情况')
    print('   - 记录异常详情')
    
    print('\n4. 添加重试机制:')
    print('   - 如果content为空，自动重试API调用')
    print('   - 设置最大重试次数')

if __name__ == '__main__':
    issues = analyze_code_issue()
    check_excel_file()
    suggest_fixes()
    
    print('\n' + '='*80)
    print('总结')
    print('='*80)
    print(f'发现 {len(issues)} 个可能导致空值的问题:')
    for i, issue in enumerate(issues, 1):
        print(f'  {i}. {issue}')
