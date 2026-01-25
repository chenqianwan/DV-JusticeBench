#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
获取最后10个案例的名称
"""
import pandas as pd
import sys

def main():
    file_path = 'data/108个案例_新标准评估_完整版_最终版.xlsx'
    
    try:
        print("正在读取文件（这可能需要一些时间）...")
        
        # 使用chunksize分块读取，但Excel不支持chunksize，所以直接读取
        # 但限制只读取需要的列
        df = pd.read_excel(
            file_path, 
            engine='openpyxl',
            usecols=['案例ID', '案例标题']  # 只读取需要的列
        )
        
        print(f"✓ 读取完成，总行数: {len(df)}")
        
        # 获取唯一案例
        case_info = df.drop_duplicates(subset=['案例ID'])
        print(f"唯一案例数: {len(case_info)}")
        
        # 按案例ID排序
        case_info = case_info.sort_values('案例ID')
        
        # 获取最后10个案例
        last_10 = case_info.tail(10)
        
        print("\n" + "=" * 80)
        print("最后10个案例:")
        print("=" * 80)
        
        for idx, (_, row) in enumerate(last_10.iterrows(), 1):
            print(f"{idx:2d}. {row['案例ID']}")
            print(f"    标题: {row['案例标题']}")
            print()
        
    except MemoryError:
        print("错误: 内存不足，尝试使用其他方法...")
        # 尝试使用openpyxl直接读取
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            
            # 读取列名
            headers = [cell.value for cell in ws[1]]
            case_id_idx = headers.index('案例ID')
            case_title_idx = headers.index('案例标题')
            
            cases = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) > max(case_id_idx, case_title_idx):
                    case_id = row[case_id_idx]
                    case_title = row[case_title_idx]
                    if case_id and case_title:
                        cases[case_id] = case_title
            
            sorted_cases = sorted(cases.items())
            last_10 = sorted_cases[-10:]
            
            print("\n" + "=" * 80)
            print("最后10个案例:")
            print("=" * 80)
            
            for idx, (case_id, case_title) in enumerate(last_10, 1):
                print(f"{idx:2d}. {case_id}")
                print(f"    标题: {case_title}")
                print()
            
            wb.close()
        except Exception as e2:
            print(f"备用方法也失败: {e2}")
            sys.exit(1)
            
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
