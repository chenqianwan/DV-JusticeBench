#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
获取最后10个案例的名称 - 轻量版本
"""
import sys

def main():
    file_path = 'data/108个案例_新标准评估_完整版_最终版.xlsx'
    
    try:
        from openpyxl import load_workbook
        print("正在读取文件...")
        
        # 使用read_only模式，只读取数据
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # 读取第一行作为列名
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        print(f"总列数: {len(headers)}")
        
        # 找到案例ID和案例标题的列索引
        try:
            case_id_idx = headers.index('案例ID')
            case_title_idx = headers.index('案例标题')
            print(f"案例ID列索引: {case_id_idx}")
            print(f"案例标题列索引: {case_title_idx}")
        except ValueError as e:
            print(f"错误: 找不到必要的列")
            print(f"可用列: {headers[:20]}")
            return
        
        # 读取所有数据，只保存唯一的案例
        cases = {}
        row_count = 0
        
        print("正在读取数据行...")
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1
            if row_count % 1000 == 0:
                print(f"  已读取 {row_count} 行...")
            
            if len(row) > max(case_id_idx, case_title_idx):
                case_id = row[case_id_idx]
                case_title = row[case_title_idx]
                
                if case_id and case_title and case_id not in cases:
                    cases[case_id] = case_title
        
        print(f"\n总行数: {row_count}")
        print(f"唯一案例数: {len(cases)}")
        
        # 按案例ID排序
        sorted_cases = sorted(cases.items())
        
        # 获取最后10个案例
        last_10 = sorted_cases[-10:]
        
        print("\n" + "=" * 80)
        print("最后10个案例:")
        print("=" * 80)
        
        for idx, (case_id, case_title) in enumerate(last_10, 1):
            print(f"{idx:2d}. {case_id}")
            print(f"    标题: {case_title}")
            print()
        
        wb.close()
        
    except ImportError:
        print("错误: 需要安装openpyxl库")
        print("请运行: pip install openpyxl")
        sys.exit(1)
    except Exception as e:
        print(f"错误: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()
