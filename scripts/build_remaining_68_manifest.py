#!/usr/bin/env python3
"""Freeze the 68 cases not used in either completed 20-case experiment."""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from typing import Any, Dict

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEMOGRAPHICS = ROOT / "data/Case_Demographics_108cases.xlsx"
QUESTIONS = ROOT / "data/108个案例_新标准评估_完整版_最终版.xlsx"
CASES_JSON = ROOT / "data/cases/cases.json"
HISTORICAL_RESULTS = ROOT / "data/results_20260112_unified_e8fd22b9/20个案例_统一评估结果_108cases.xlsx"
ADDITIONAL_20 = ROOT / "data/additional_20_cases_20260819.json"
OUTPUT = ROOT / "data/remaining_68_cases_20260821.json"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def worksheet_records(path: Path, sheet_name: str) -> list[Dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(rows)]
    return [dict(zip(headers, row)) for row in rows]


def main() -> int:
    population_rows = worksheet_records(DEMOGRAPHICS, "108案例列表")
    population_ids = [str(row["案例ID"]) for row in population_rows]
    if len(population_ids) != 108 or len(set(population_ids)) != 108:
        raise ValueError("population must contain exactly 108 unique cases")

    historical_rows = worksheet_records(HISTORICAL_RESULTS, "GPT-4o")
    historical_ids = {str(row["案例ID"]) for row in historical_rows}
    additional_ids = {str(item["case_id"]) for item in load_json(ADDITIONAL_20)["cases"]}
    if len(historical_ids) != 20 or len(additional_ids) != 20:
        raise ValueError("completed experiments must each contain exactly 20 cases")
    if historical_ids & additional_ids:
        raise ValueError("the two completed 20-case sets unexpectedly overlap")

    excluded_ids = historical_ids | additional_ids
    remaining_ids = [case_id for case_id in population_ids if case_id not in excluded_ids]
    if len(remaining_ids) != 68:
        raise ValueError(f"expected 68 remaining cases, found {len(remaining_ids)}")

    question_rows = worksheet_records(QUESTIONS, "Sheet1")
    questions_by_case: Dict[str, list[str]] = {}
    for row in question_rows:
        case_id = str(row["案例ID"])
        question = str(row["问题"] or "").strip()
        if question:
            questions_by_case.setdefault(case_id, []).append(question)

    case_payloads = load_json(CASES_JSON)
    population_by_id = {str(row["案例ID"]): row for row in population_rows}
    manifest_cases = []
    for case_id in remaining_ids:
        questions = questions_by_case.get(case_id) or []
        payload = case_payloads.get(case_id) or {}
        case_text = str(payload.get("content", payload.get("case_text", "")) or "")
        judgment = str(payload.get("judge_decision") or "")
        if len(questions) != 5 or len(set(questions)) != 5:
            raise ValueError(f"{case_id} does not have exactly five unique non-empty questions")
        if not case_text or not judgment:
            raise ValueError(f"{case_id} is missing case text or judgment")
        row = population_by_id[case_id]
        manifest_cases.append(
            {
                "case_id": case_id,
                "title": str(row["案例标题"] or payload.get("title") or ""),
                "cause": str(row["案由"] or "其他"),
                "question_count": 5,
                "case_text_chars": len(case_text),
                "judgment_chars": len(judgment),
            }
        )

    manifest = {
        "manifest_version": "1.0",
        "created_at": date.today().isoformat(),
        "purpose": "All 68 cases not included in the historical or additional 20-case experiments",
        "population_case_count": 108,
        "excluded_historical_case_count": 20,
        "excluded_additional_case_count": 20,
        "selected_case_count": 68,
        "selection_method": "Population order after exact case_id exclusion of both completed 20-case sets; no sampling.",
        "sources": {
            "population": "data/Case_Demographics_108cases.xlsx#108案例列表",
            "historical_completed_set": "data/results_20260112_unified_e8fd22b9/20个案例_统一评估结果_108cases.xlsx#GPT-4o",
            "additional_completed_set": "data/additional_20_cases_20260819.json",
            "questions": "data/108个案例_新标准评估_完整版_最终版.xlsx#Sheet1",
            "case_text_and_judgment": "data/cases/cases.json",
        },
        "validation": {
            "overlap_with_historical_20": 0,
            "overlap_with_additional_20": 0,
            "duplicate_case_ids": 0,
            "cases_with_exactly_five_unique_nonempty_questions": 68,
            "cases_with_nonempty_case_text": 68,
            "cases_with_nonempty_judgment": 68,
        },
        "cases": manifest_cases,
    }
    OUTPUT.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"manifest={OUTPUT}")
    print("cases=68 excluded=40 population=108")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
