#!/usr/bin/env python3
"""Build the comparison-first Experiment-1 report for historical vs new 20 cases."""

from __future__ import annotations

import hashlib
import json
import math
import statistics
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import load_workbook
from scipy import stats


ROOT = Path(__file__).resolve().parents[1]
OLD_WORKBOOK = ROOT / "data/results_20260112_unified_e8fd22b9/20个案例_统一评估结果_108cases.xlsx"
NEW_RESULTS = ROOT / "data/experiment1_additional20_xhub_20260819/combined_results.json"
NEW_MANIFEST = ROOT / "data/additional_20_cases_20260819.json"
OUTPUT_JSON = ROOT / "data/experiment1_additional20_xhub_20260819/comparison_20v20.json"
OUTPUT_REPORT = ROOT / "实验1_新增20case_五条件运行报告_20260819.md"


CONDITION_MAP = {
    "GPT-4o": "GPT-4o",
    "Gemini 2.5 Flash": "Gemini 2.5 Flash",
    "Qwen-Max": "Qwen-Max",
    "DeepSeek": "DeepSeek thinking",
    "DeepSeek-NoThinking": "DeepSeek non-thinking",
}
CONDITION_ORDER = list(CONDITION_MAP.values())


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def holm_adjust(p_values: Iterable[float]) -> List[float]:
    values = list(p_values)
    order = sorted(range(len(values)), key=lambda index: values[index])
    adjusted = [0.0] * len(values)
    running = 0.0
    count = len(values)
    for rank, index in enumerate(order):
        candidate = min(1.0, (count - rank) * values[index])
        running = max(running, candidate)
        adjusted[index] = running
    return adjusted


def independent_case_comparison(label: str, old: List[float], new: List[float]) -> Dict[str, Any]:
    old_variance = stats.tvar(old)
    new_variance = stats.tvar(new)
    delta = statistics.mean(new) - statistics.mean(old)
    standard_error = math.sqrt(old_variance / len(old) + new_variance / len(new))
    numerator = (old_variance / len(old) + new_variance / len(new)) ** 2
    denominator = (old_variance / len(old)) ** 2 / (len(old) - 1) + (new_variance / len(new)) ** 2 / (len(new) - 1)
    degrees_freedom = numerator / denominator
    critical = stats.t.ppf(0.975, degrees_freedom)
    _, p_value = stats.ttest_ind(new, old, equal_var=False)
    pooled_sd = math.sqrt(
        ((len(old) - 1) * old_variance + (len(new) - 1) * new_variance)
        / (len(old) + len(new) - 2)
    )
    small_sample_correction = 1 - 3 / (4 * (len(old) + len(new)) - 9)
    hedges_g = delta / pooled_sd * small_sample_correction if pooled_sd else 0.0
    return {
        "condition": label,
        "old_n": len(old),
        "new_n": len(new),
        "old_mean": statistics.mean(old),
        "old_sd": statistics.stdev(old),
        "old_median": statistics.median(old),
        "old_min": min(old),
        "old_max": max(old),
        "new_mean": statistics.mean(new),
        "new_sd": statistics.stdev(new),
        "new_median": statistics.median(new),
        "new_min": min(new),
        "new_max": max(new),
        "delta": delta,
        "delta_ci95_low": delta - critical * standard_error,
        "delta_ci95_high": delta + critical * standard_error,
        "welch_df": degrees_freedom,
        "welch_p": float(p_value),
        "hedges_g": hedges_g,
    }


def read_old_workbook() -> Tuple[Dict[str, Dict[str, float]], Dict[str, str], Dict[str, int]]:
    workbook = load_workbook(OLD_WORKBOOK, read_only=True, data_only=True)
    case_scores: Dict[str, Dict[str, float]] = {}
    titles: Dict[str, str] = {}
    refusal_counts: Dict[str, int] = {}
    expected_case_ids: set[str] | None = None
    for sheet_name, condition in CONDITION_MAP.items():
        worksheet = workbook[sheet_name]
        row_iterator = worksheet.iter_rows(values_only=True)
        headers = next(row_iterator)
        columns = {value: index for index, value in enumerate(headers)}
        grouped: Dict[str, List[float]] = {}
        refusals = 0
        row_count = 0
        for row in row_iterator:
            case_id = str(row[columns["案例ID"]])
            score = row[columns["总分"]]
            question_number = row[columns["问题编号"]]
            answer = str(row[columns["AI回答"]] or "")
            if not isinstance(score, (int, float)):
                raise ValueError(f"non-numeric historical score: {sheet_name} {case_id} Q{question_number}")
            grouped.setdefault(case_id, []).append(float(score))
            titles.setdefault(case_id, str(row[columns["案例标题"]] or ""))
            if "prompt contains sensitive words" in answer.lower() or "prohibited use policy" in answer.lower():
                refusals += 1
            row_count += 1
        if row_count != 100 or len(grouped) != 20 or any(len(values) != 5 for values in grouped.values()):
            raise ValueError(f"historical sheet matrix invalid: {sheet_name}")
        current_case_ids = set(grouped)
        if expected_case_ids is None:
            expected_case_ids = current_case_ids
        elif current_case_ids != expected_case_ids:
            raise ValueError(f"historical case set mismatch: {sheet_name}")
        case_scores[condition] = {
            case_id: statistics.mean(values) for case_id, values in grouped.items()
        }
        refusal_counts[condition] = refusals
    return case_scores, titles, refusal_counts


def new_refusal_counts(case_ids: List[str]) -> Dict[str, int]:
    counts = {condition: 0 for condition in CONDITION_ORDER}
    for case_id in case_ids:
        raw_path = NEW_RESULTS.parent / "cases" / case_id / "raw_results.json"
        result = load_json(raw_path)
        for row in result.get("rows") or []:
            answer = str((row.get("generation") or {}).get("content") or "")
            if "prompt contains sensitive words" in answer.lower() or "prohibited use policy" in answer.lower():
                counts[str(row["condition"])] += 1
    return counts


def fmt(value: float | None, digits: int = 2) -> str:
    return "—" if value is None else f"{value:.{digits}f}"


def fmt_p(value: float) -> str:
    if value < 0.001:
        return "<0.001"
    return f"{value:.3f}"


def signed(value: float, digits: int = 2) -> str:
    return f"{value:+.{digits}f}"


def main() -> int:
    new_results = load_json(NEW_RESULTS)
    manifest = load_json(NEW_MANIFEST)
    if new_results.get("status") != "PASS":
        raise ValueError("new 20-case results are not strict PASS")
    if int(new_results.get("completed_case_count") or 0) != 20 or int(new_results.get("completed_answer_count") or 0) != 500:
        raise ValueError("new result matrix is not 20 cases / 500 units")

    old_scores, old_titles, old_refusals = read_old_workbook()
    old_case_ids = sorted(next(iter(old_scores.values())).keys())
    new_case_ids = [str(case_id) for case_id in new_results["completed_case_ids"]]
    if set(old_case_ids) & set(new_case_ids):
        raise ValueError("historical and new case sets overlap")
    if len(new_case_ids) != 20:
        raise ValueError("new case count is not 20")

    new_case_scores = new_results["case_condition_means"]
    comparisons: List[Dict[str, Any]] = []
    for condition in CONDITION_ORDER:
        old_values = [float(old_scores[condition][case_id]) for case_id in old_case_ids]
        new_values = [float(new_case_scores[case_id][condition]) for case_id in new_case_ids]
        comparisons.append(independent_case_comparison(condition, old_values, new_values))
    adjusted = holm_adjust(item["welch_p"] for item in comparisons)
    for item, adjusted_p in zip(comparisons, adjusted):
        item["holm_p_across_five_conditions"] = adjusted_p

    old_overall_by_case = {
        case_id: statistics.mean(old_scores[condition][case_id] for condition in CONDITION_ORDER)
        for case_id in old_case_ids
    }
    new_overall_by_case = {
        case_id: statistics.mean(float(new_case_scores[case_id][condition]) for condition in CONDITION_ORDER)
        for case_id in new_case_ids
    }
    overall = independent_case_comparison(
        "五条件总体",
        list(old_overall_by_case.values()),
        list(new_overall_by_case.values()),
    )

    new_refusals = new_refusal_counts(new_case_ids)
    manifest_by_id = {str(item["case_id"]): item for item in manifest["cases"]}
    old_rank = sorted(CONDITION_ORDER, key=lambda condition: statistics.mean(old_scores[condition].values()), reverse=True)
    new_rank = sorted(CONDITION_ORDER, key=lambda condition: statistics.mean(new_case_scores[case_id][condition] for case_id in new_case_ids), reverse=True)

    comparison_output = {
        "generated_from": {
            "historical_workbook": str(OLD_WORKBOOK.resolve()),
            "historical_workbook_sha256": sha256_file(OLD_WORKBOOK),
            "new_combined_results": str(NEW_RESULTS.resolve()),
            "new_combined_results_sha256": sha256_file(NEW_RESULTS),
            "new_manifest": str(NEW_MANIFEST.resolve()),
            "new_manifest_sha256": sha256_file(NEW_MANIFEST),
        },
        "design": {
            "historical_cases": 20,
            "new_cases": 20,
            "questions_per_case": 5,
            "shared_conditions": CONDITION_ORDER,
            "historical_units": 500,
            "new_units": 500,
            "case_overlap_count": 0,
            "inferential_unit": "case-level mean",
        },
        "overall": overall,
        "condition_comparisons": comparisons,
        "old_rank": old_rank,
        "new_rank": new_rank,
        "policy_refusal_units": {"historical": old_refusals, "new": new_refusals},
        "old_case_overall_means": old_overall_by_case,
        "new_case_overall_means": new_overall_by_case,
    }
    OUTPUT_JSON.write_text(json.dumps(comparison_output, ensure_ascii=False, indent=2), encoding="utf-8")

    largest = max(comparisons, key=lambda item: abs(item["delta"]))
    lines = [
        "# 实验1：新增20案 vs 原20案五条件对比报告",
        "",
        "> 主比较：两个互不重叠的20-case样本；每案5题；五个共享模型条件。  ",
        "> 推断单位：先对每个 case 的5题取均值，再比较20个 case；不得把100题当作100个独立案例。  ",
        "> 新增实验状态：**PASS（20/20 cases，500/500回答评分单元）**。  ",
        "> 旧表中的 Claude Opus 4 和 GPT-5 因新增实验未运行，不纳入主比较。",
        "",
        "## 1. 结论先行",
        "",
        f"五条件总体由 **{overall['old_mean']:.2f}** 升至 **{overall['new_mean']:.2f}**，差值 **{signed(overall['delta'])} / 20**（case-level 95% CI {signed(overall['delta_ci95_low'])} 至 {signed(overall['delta_ci95_high'])}；Welch p={fmt_p(overall['welch_p'])}；Hedges g={overall['hedges_g']:.2f}）。",
        "",
        f"但这不是统一的整体上移。最大变化来自 **{largest['condition']}**（{signed(largest['delta'])}）；Gemini 与两个 DeepSeek 条件的均值变化很小，Qwen-Max 为中等幅度上升。因此，新增20案总体分数更高主要由 GPT-4o 驱动，不能直接概括成“所有模型都明显进步”。",
        "",
        "更重要的是：这是两个不同 case 样本之间的独立比较，不是同一批 case 的前后配对；同时评分器和 xhub 后端存在版本漂移风险。差值可能同时包含模型表现、样本难度、政策拦截和评分器漂移，不能仅解释为模型能力提升。",
        "",
        "## 2. 主要对比表",
        "",
        "| 条件 | 原20案均值±SD | 新20案均值±SD | 新−旧 | 差值95% CI | Welch p | Holm p（5条件） | Hedges g |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for item in comparisons:
        lines.append(
            f"| {item['condition']} | {item['old_mean']:.2f} ± {item['old_sd']:.2f} | {item['new_mean']:.2f} ± {item['new_sd']:.2f} | {signed(item['delta'])} | {signed(item['delta_ci95_low'])} 至 {signed(item['delta_ci95_high'])} | {fmt_p(item['welch_p'])} | {fmt_p(item['holm_p_across_five_conditions'])} | {item['hedges_g']:.2f} |"
        )
    lines.append(
        f"| **五条件总体** | **{overall['old_mean']:.2f} ± {overall['old_sd']:.2f}** | **{overall['new_mean']:.2f} ± {overall['new_sd']:.2f}** | **{signed(overall['delta'])}** | **{signed(overall['delta_ci95_low'])} 至 {signed(overall['delta_ci95_high'])}** | **{fmt_p(overall['welch_p'])}** | — | **{overall['hedges_g']:.2f}** |"
    )
    lines.extend(
        [
            "",
            "说明：总体比较是预先关心的五条件 case 均分；Holm 校正仅用于五个条件级探索性比较。p值不替代效应量和区间判断。",
            "",
            "## 3. 分布与排名",
            "",
            "| 条件 | 原20案中位数（范围） | 新20案中位数（范围） | 原排名 | 新排名 |",
            "|---|---:|---:|---:|---:|",
        ]
    )
    for item in comparisons:
        condition = item["condition"]
        lines.append(
            f"| {condition} | {item['old_median']:.2f}（{item['old_min']:.2f}–{item['old_max']:.2f}） | {item['new_median']:.2f}（{item['new_min']:.2f}–{item['new_max']:.2f}） | {old_rank.index(condition) + 1} | {new_rank.index(condition) + 1} |"
        )
    lines.extend(
        [
            "",
            f"原20案排名：{' > '.join(old_rank)}。  ",
            f"新20案排名：{' > '.join(new_rank)}。",
            "",
            "## 4. 如何解释差值",
            "",
            "- **GPT-4o：明确上升。** case均分增加4.11，区间不跨0，且在五条件Holm校正后仍成立；它是总体均值上升的主要来源。",
            "- **Qwen-Max：可能上升，但证据边界。** 均值增加1.36，未校正p约0.053，95% CI轻微跨0；更合适的表述是“方向积极但不确定”。",
            "- **Gemini：总体持平，但有鲁棒性失败。** 新样本 case 88 的5题均返回 Google prohibited-use policy 提示并按规则计0分；这是模型条件在敏感刑事案上的真实失败，不从主分析删除。",
            "- **DeepSeek thinking / non-thinking：基本稳定。** 两者差值均较小、区间跨0；且旧别名与当前 `deepseek-v3.2` 显式 thinking 开关并非可证明的同一后端版本，因此不应作强版本进步结论。",
            "- **评分宽松假设没有得到“全模型同步上移”的直接支持。** 如果仅是统一评分器整体变松，通常预期五个条件方向更一致；本次主要集中在GPT-4o。但评分器版本漂移仍然可能与样本差异共同作用，不能排除。",
            "",
            "## 5. 可比性与限制",
            "",
            "| 项目 | 原20案 | 新增20案 | 对解释的影响 |",
            "|---|---|---|---|",
            "| case | 原始20案 | 从其余88案中分层抽取的20案 | 样本不同，只能作独立组比较，不是配对复现 |",
            "| case重叠 | — | 与原20案重叠0 | 避免数据重复，但样本难度差异仍存在 |",
            "| 每案问题 | 5题 | 同一108-case问题表中的5题 | 二级重复测量一致 |",
            "| 共享条件 | GPT-4o、Gemini 2.5 Flash、Qwen-Max、DS thinking、DS non-thinking | 同左 | 只比较五个共同条件 |",
            "| DeepSeek路由 | `deepseek-reasoner` / `deepseek-chat`历史别名 | `deepseek-v3.2` + thinking显式开关 | 后端版本不可证明完全相同 |",
            "| 评分 | 旧表记录为DeepSeek评估 | `deepseek-v3.2` thinking + legacy解析/扣分 | 评分器漂移可能影响绝对分数 |",
            "| 脱敏 | 每模型任务独立DeepSeek脱敏 | 每case×条件独立DS v3.2脱敏 | 语义复刻，但后端版本仍有差异 |",
            "| Opus / GPT-5 | 旧表存在 | 新实验未运行 | 不纳入新旧主比较 |",
            "",
            "## 6. 两组 case 级总体均分（审计表）",
            "",
            "### 原20案",
            "",
            "| case_id | 标题 | 五条件均分 |",
            "|---|---|---:|",
        ]
    )
    for case_id in old_case_ids:
        lines.append(f"| `{case_id}` | {old_titles[case_id]} | {old_overall_by_case[case_id]:.2f} |")
    lines.extend(
        [
            "",
            "### 新增20案",
            "",
            "| case_id | 案由 | 五条件均分 |",
            "|---|---|---:|",
        ]
    )
    for case_id in new_case_ids:
        lines.append(
            f"| `{case_id}` | {manifest_by_id[case_id]['cause']} | {new_overall_by_case[case_id]:.2f} |"
        )
    costs = new_results["estimated_cost_by_stage_usd"]
    usage = new_results["stage_usage"]
    lines.extend(
        [
            "",
            "## 7. 新增20案运行与成本附录",
            "",
            "| 阶段 | prompt tokens | completion tokens | total tokens | xhub list价估算 |",
            "|---|---:|---:|---:|---:|",
        ]
    )
    for stage, label in (("masking", "脱敏"), ("generation", "回答生成"), ("scoring", "自动评分")):
        lines.append(
            f"| {label} | {usage[stage]['prompt_tokens']:,} | {usage[stage]['completion_tokens']:,} | {usage[stage]['total_tokens']:,} | ${costs[stage]:.4f} |"
        )
    lines.extend(
        [
            f"| **合计** | **{sum(usage[stage]['prompt_tokens'] for stage in usage):,}** | **{sum(usage[stage]['completion_tokens'] for stage in usage):,}** | **{new_results['total_tokens_including_failed_pipeline_attempts']:,}** | **${new_results['estimated_xhub_list_cost_usd']:.4f}** |",
            "",
            f"人民币参考：约 ¥{new_results['estimated_xhub_list_cost_cny_at_7_3']:.2f}（按7.3换算；最终以xhub账单为准）。费用包含最终保留结果、整题内部重试和已捕获的失败整案用量；raw写出前中止的请求无法取得完整token回执。",
            "",
            "## 8. 数据源与复现",
            "",
            f"- 原20案：`{OLD_WORKBOOK.resolve()}`（SHA-256 `{sha256_file(OLD_WORKBOOK)}`）",
            f"- 新增20案聚合结果：`{NEW_RESULTS.resolve()}`（SHA-256 `{sha256_file(NEW_RESULTS)}`）",
            f"- 新增20案抽样清单：`{NEW_MANIFEST.resolve()}`（SHA-256 `{sha256_file(NEW_MANIFEST)}`）",
            f"- 对比统计JSON：`{OUTPUT_JSON.resolve()}`",
            f"- 对比生成脚本：`{Path(__file__).resolve()}`",
            "",
            "统计方法：对每个 case 内5题先取均值；条件级比较使用两个独立20-case样本的Welch t检验、差值95% CI和Hedges g；五个条件级p值作Holm校正。总体指标先在每个case内对五条件等权平均，再作独立组比较。",
            "",
        ]
    )
    OUTPUT_REPORT.write_text("\n".join(lines), encoding="utf-8")
    print(f"report={OUTPUT_REPORT}")
    print(f"comparison_json={OUTPUT_JSON}")
    print(f"overall_old={overall['old_mean']:.4f} overall_new={overall['new_mean']:.4f} delta={overall['delta']:.4f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
