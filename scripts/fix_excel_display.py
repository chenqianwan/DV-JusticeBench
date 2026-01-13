#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
修复Excel文件中AI回答列的显示问题
确保所有内容都可见
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.colors import Color

def fix_ai_answer_display(file_path, case_id=None):
    """修复AI回答列的显示问题"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # 找到AI回答列（J列）
    ai_answer_col = 'J'
    
    print(f'处理文件: {file_path}')
    print(f'总行数: {ws.max_row}')
    
    fixed_count = 0
    
    # 遍历所有数据行（从第2行开始，第1行是表头）
    for row_num in range(2, ws.max_row + 1):
        cell = ws[f'{ai_answer_col}{row_num}']
        
        # 如果指定了案例ID，只处理该案例
        if case_id:
            case_id_cell = ws[f'A{row_num}']
            if case_id_cell.value != case_id:
                continue
        
        # 检查单元格是否有值但可能不可见
        if cell.value is not None and str(cell.value).strip() != '':
            # 确保字体颜色可见（黑色）
            if cell.font is None or cell.font.color is None:
                cell.font = Font(name='Calibri', size=11, color=Color(rgb='00000000'))
            elif hasattr(cell.font.color, 'rgb') and cell.font.color.rgb == 'FFFFFFFF':
                # 如果是白色字体，改为黑色
                cell.font = Font(name='Calibri', size=11, color=Color(rgb='00000000'))
            
            # 确保文本对齐
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # 确保行高足够显示内容
            if ws.row_dimensions[row_num].height is None or ws.row_dimensions[row_num].height < 15:
                # 根据内容长度设置行高
                content_length = len(str(cell.value))
                estimated_height = max(15, min(60, content_length / 50 * 15))
                ws.row_dimensions[row_num].height = estimated_height
            
            fixed_count += 1
    
    # 确保列宽足够
    if ws.column_dimensions[ai_answer_col].width is None or ws.column_dimensions[ai_answer_col].width < 30:
        ws.column_dimensions[ai_answer_col].width = 50
    
    # 保存文件
    output_path = file_path.replace('.xlsx', '_fixed.xlsx')
    wb.save(output_path)
    wb.close()
    
    print(f'✓ 修复完成，已保存到: {output_path}')
    print(f'  修复了 {fixed_count} 个单元格')
    
    return output_path

if __name__ == '__main__':
    import sys
    
    file_path = 'data/results_20260112_105927/5个案例_统一评估结果_108cases.xlsx'
    case_id = 'case_20251230_101232_6' if len(sys.argv) < 2 else None
    
    fix_ai_answer_display(file_path, case_id)
