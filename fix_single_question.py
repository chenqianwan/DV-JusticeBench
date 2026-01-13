#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
单独修复一个问题：重新生成AI回答并更新Excel
"""
import pandas as pd
import openpyxl
from datetime import datetime
import os
import sys

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from utils.ai_api import UnifiedAIAPI
from utils.evaluator import AnswerEvaluator
from utils.data_masking import DataMaskerAPI

def fix_single_question(case_id, question_num, excel_file, sheet_name='DeepSeek'):
    """
    修复单个问题的AI回答
    
    Args:
        case_id: 案例ID
        question_num: 问题编号
        excel_file: Excel文件路径
        sheet_name: 工作表名称
    """
    print('='*80)
    print(f'修复问题：案例ID={case_id}, 问题编号={question_num}')
    print('='*80)
    
    # 读取Excel文件
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[sheet_name]
    
    # 找到对应的行
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if (ws[f'A{row_idx}'].value == case_id and 
            ws[f'D{row_idx}'].value == question_num):
            target_row = row_idx
            break
    
    if not target_row:
        print(f"错误：未找到案例ID={case_id}, 问题编号={question_num}的行")
        wb.close()
        return False
    
    print(f"找到目标行：{target_row}")
    
    # 读取该行的数据
    row_data = {
        '案例ID': ws[f'A{target_row}'].value,
        '案例标题': ws[f'B{target_row}'].value,
        '案例标题（脱敏）': ws[f'C{target_row}'].value,
        '问题编号': ws[f'D{target_row}'].value,
        '问题': ws[f'E{target_row}'].value,
        '使用的模型': ws[f'F{target_row}'].value,
    }
    
    print(f"\n问题：{row_data['问题'][:80]}...")
    
    # 需要从统一数据中获取脱敏内容和法官判决
    # 查找统一数据文件
    unified_file = None
    for root, dirs, files in os.walk('data'):
        for file in files:
            if '统一评估结果' in file and file.endswith('.xlsx') and not file.startswith('.~'):
                # 检查是否是统一数据文件（通常包含所有案例的问题和脱敏数据）
                full_path = os.path.join(root, file)
                try:
                    df = pd.read_excel(full_path, sheet_name='DeepSeek')
                    if len(df) > 0 and '案例ID' in df.columns:
                        # 检查是否包含该案例
                        case_df = df[df['案例ID'] == case_id]
                        if len(case_df) > 0:
                            unified_file = full_path
                            print(f"找到统一数据文件: {unified_file}")
                            break
                except:
                    pass
        if unified_file:
            break
    
    if not unified_file:
        print("警告：未找到统一数据文件，将尝试从当前文件提取数据")
        # 从当前文件的其他行获取脱敏数据（假设同一案例的其他问题有数据）
        for row_idx in range(2, ws.max_row + 1):
            if ws[f'A{row_idx}'].value == case_id and row_idx != target_row:
                # 使用其他行的脱敏数据（假设同一案例的脱敏数据相同）
                masked_title = ws[f'C{row_idx}'].value
                break
    else:
        # 从统一数据文件读取
        df = pd.read_excel(unified_file, sheet_name='DeepSeek')
        case_df = df[df['案例ID'] == case_id]
        if len(case_df) > 0:
            first_row = case_df.iloc[0]
            masked_title = first_row.get('案例标题（脱敏）', '')
    
    # 需要获取案例的脱敏内容和法官判决
    # 从cases.json加载案例数据
    import json
    cases_json_file = 'data/cases/cases.json'
    
    if not os.path.exists(cases_json_file):
        print(f"错误：未找到案例数据文件 {cases_json_file}")
        wb.close()
        return False
    
    with open(cases_json_file, 'r', encoding='utf-8') as f:
        cases_data = json.load(f)
    
    if case_id not in cases_data:
        print(f"错误：未找到案例 {case_id} 的原始数据")
        wb.close()
        return False
    
    case = cases_data[case_id]
    
    # 执行脱敏
    print("\n执行脱敏处理...")
    masker = DataMaskerAPI()
    case_dict = {
        'title': case['title'],
        'case_text': case.get('content', case.get('case_text', '')),
        'judge_decision': case.get('judge_decision', '')
    }
    
    masked_case = masker.mask_case_with_api(case_dict)
    masked_content = masked_case.get('case_text_masked', '')
    masked_judge = masked_case.get('judge_decision_masked', '')
    
    print(f"脱敏完成，内容长度: {len(masked_content)} 字符")
    
    # 生成AI回答
    print(f"\n生成AI回答（使用DeepSeek，thinking模式）...")
    deepseek_api = UnifiedAIAPI(provider='deepseek')
    ai_response = deepseek_api.analyze_case(masked_content, question=row_data['问题'], use_thinking=True)
    
    if isinstance(ai_response, dict):
        ai_answer = ai_response.get('answer', '')
        ai_thinking = ai_response.get('thinking', '')
    else:
        ai_answer = ai_response
        ai_thinking = ''
    
    if not ai_answer or not ai_answer.strip():
        print(f"错误：AI回答仍为空")
        wb.close()
        return False
    
    print(f"✓ AI回答生成完成（{len(ai_answer)}字符）")
    
    # 进行评估（使用现有的评估逻辑，不修改评估代码）
    print(f"\n进行评估...")
    evaluator = AnswerEvaluator()
    evaluation = evaluator.evaluate_answer(
        ai_answer=ai_answer,
        judge_decision=masked_judge,
        question=row_data['问题'],
        case_text=masked_content
    )
    
    print(f"✓ 评估完成（总分: {evaluation['总分']:.2f}/20）")
    
    # 更新Excel文件
    print(f"\n更新Excel文件...")
    
    # 找到J列（AI回答）和K列（AI回答Thinking）
    j_col = None
    k_col = None
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header == 'AI回答':
            j_col = col_idx
        elif header == 'AI回答Thinking':
            k_col = col_idx
    
    if j_col:
        ws.cell(row=target_row, column=j_col).value = ai_answer
        print(f"  ✓ 更新AI回答（列{j_col}）")
    
    if k_col:
        ws.cell(row=target_row, column=k_col).value = ai_thinking if ai_thinking else ''
        print(f"  ✓ 更新AI回答Thinking（列{k_col}）")
    
    # 更新评估结果
    # 找到总分列
    score_col = None
    percent_col = None
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header == '总分':
            score_col = col_idx
        elif header == '百分制':
            percent_col = col_idx
    
    if score_col:
        ws.cell(row=target_row, column=score_col).value = evaluation['总分']
    if percent_col:
        ws.cell(row=target_row, column=percent_col).value = evaluation['百分制']
    
    # 更新其他评估字段
    dimension_cols = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header and '_得分' in str(header):
            dimension_cols[header] = col_idx
    
    dimension_scores = evaluation.get('各维度得分', {})
    for dim_name, score in dimension_scores.items():
        col_name = f'{dim_name}_得分'
        if col_name in dimension_cols:
            ws.cell(row=target_row, column=dimension_cols[col_name]).value = score
    
    # 更新错误和评价字段
    error_cols = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header in ['详细评价', '评价Thinking', '错误标记', '重大错误', '明显错误', '微小错误']:
            error_cols[header] = col_idx
    
    if '详细评价' in error_cols:
        ws.cell(row=target_row, column=error_cols['详细评价']).value = evaluation.get('详细评价', '')
    if '评价Thinking' in error_cols:
        ws.cell(row=target_row, column=error_cols['评价Thinking']).value = evaluation.get('评价Thinking', '')
    
    error_details = evaluation.get('错误详情', {})
    if '重大错误' in error_cols:
        ws.cell(row=target_row, column=error_cols['重大错误']).value = '; '.join(error_details.get('重大错误', [])) if error_details.get('重大错误') else None
    if '明显错误' in error_cols:
        ws.cell(row=target_row, column=error_cols['明显错误']).value = '; '.join(error_details.get('明显错误', [])) if error_details.get('明显错误') else None
    if '微小错误' in error_cols:
        ws.cell(row=target_row, column=error_cols['微小错误']).value = '; '.join(error_details.get('微小错误', [])) if error_details.get('微小错误') else None
    
    # 清除处理错误字段（如果有）
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header == '处理错误':
            ws.cell(row=target_row, column=col_idx).value = ''
            break
    
    # 保存文件
    output_file = excel_file.replace('.xlsx', '_fixed.xlsx')
    wb.save(output_file)
    wb.close()
    
    print(f"\n✓ 修复完成，已保存到: {output_file}")
    print(f"  原文件: {excel_file}")
    print(f"  新文件: {output_file}")
    
    return True

if __name__ == '__main__':
    case_id = 'case_20251230_101232_6'
    question_num = 1
    excel_file = 'data/results_20260112_105927/5个案例_统一评估结果_108cases.xlsx'
    sheet_name = 'DeepSeek'
    
    success = fix_single_question(case_id, question_num, excel_file, sheet_name)
    
    if success:
        print("\n" + "="*80)
        print("修复成功！")
        print("="*80)
    else:
        print("\n" + "="*80)
        print("修复失败，请查看上面的错误信息")
        print("="*80)
        sys.exit(1)
