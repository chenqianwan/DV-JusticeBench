"""
Excel导出模块
使用pandas生成包含案例、问题、AI结果、法官判决和差异对比的Excel文件
"""
import pandas as pd
import os
from datetime import datetime
from typing import List, Dict
from config import RESULTS_DIR
from utils.data_masking import DataMaskerAPI


class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self, results_dir: str = None):
        """
        初始化导出器
        
        Args:
            results_dir: 结果存储目录，如果不提供则从config读取
        """
        self.results_dir = results_dir or RESULTS_DIR
        os.makedirs(self.results_dir, exist_ok=True)
        self.data_masker = DataMaskerAPI()  # 初始化API脱敏工具
    
    def export_analysis_results(self, results: List[Dict], filename: str = None) -> str:
        """
        导出分析结果到Excel
        
        Args:
            results: 分析结果列表，每个结果包含：
                - case_id: 案例ID
                - case_title: 案例标题
                - case_text: 案例文本
                - question: 问题
                - ai_decision: AI判决
                - judge_decision: 法官判决
                - comparison: 差异对比（可选）
            filename: 输出文件名，如果不提供则自动生成
            
        Returns:
            生成的Excel文件路径
        """
        if not results:
            raise ValueError("结果列表为空")
        
        # 准备数据
        data = []
        for result in results:
            row = {
                '案例ID': result.get('case_id', ''),
                '案例标题': result.get('case_title', ''),
                '案例内容': result.get('case_text', ''),
                '问题': result.get('question', ''),
                'AI判决': result.get('ai_decision', ''),
                '法官判决': result.get('judge_decision', ''),
                '差异对比': result.get('comparison', ''),
                '生成时间': result.get('created_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            }
            data.append(row)
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'analysis_results_{timestamp}.xlsx'
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.results_dir, filename)
        
        # 导出到Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='分析结果', index=False)
            
            # 获取工作表对象以调整列宽
            worksheet = writer.sheets['分析结果']
            
            # 自动调整列宽
            from openpyxl.utils import get_column_letter
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                )
                # 设置最大宽度限制
                adjusted_width = min(max_length + 2, 100)
                column_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return filepath
    
    def export_case_summary(self, cases: List[Dict], filename: str = None) -> str:
        """
        导出案例摘要到Excel
        
        Args:
            cases: 案例列表
            filename: 输出文件名，如果不提供则自动生成
            
        Returns:
            生成的Excel文件路径
        """
        if not cases:
            raise ValueError("案例列表为空")
        
        # 准备数据
        data = []
        for case in cases:
            row = {
                '案例ID': case.get('id', ''),
                '案例标题': case.get('title', ''),
                '案例日期': case.get('case_date', ''),
                '案例内容': case.get('case_text', '')[:500] + '...' if len(case.get('case_text', '')) > 500 else case.get('case_text', ''),
                '法官判决': case.get('judge_decision', '')[:500] + '...' if len(case.get('judge_decision', '')) > 500 else case.get('judge_decision', ''),
                '创建时间': case.get('created_at', '')
            }
            data.append(row)
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'case_summary_{timestamp}.xlsx'
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.results_dir, filename)
        
        # 导出到Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='案例摘要', index=False)
            
            # 获取工作表对象以调整列宽
            worksheet = writer.sheets['案例摘要']
            
            # 自动调整列宽
            from openpyxl.utils import get_column_letter
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                )
                adjusted_width = min(max_length + 2, 100)
                column_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return filepath
    
    def generate_import_template(self, filename: str = None) -> str:
        """
        生成Excel导入模板
        
        Args:
            filename: 输出文件名，如果不提供则自动生成
            
        Returns:
            生成的Excel文件路径
        """
        # 创建模板数据（包含示例行）
        data = [
            {
                '案例标题': '示例案例1：故意伤害案',
                '案例内容': '被告人张三因与李四发生口角，持刀将李四砍伤，致其轻伤二级。',
                '案例日期': '2024-01-15',
                '法官判决': '法院认为，被告人张三故意伤害他人身体，致人轻伤，其行为已构成故意伤害罪。鉴于其认罪态度较好，判处有期徒刑一年，缓刑一年。'
            },
            {
                '案例标题': '示例案例2：合同纠纷案',
                '案例内容': '原告王五与被告赵六签订房屋买卖合同，后因房价上涨，被告拒绝履行合同。',
                '案例日期': '2024-02-20',
                '法官判决': ''
            }
        ]
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 生成文件名
        if not filename:
            filename = '案例导入模板.xlsx'
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # 使用临时目录存储模板文件
        import tempfile
        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, filename)
        
        # 导出到Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='案例数据', index=False)
            
            # 获取工作表对象以调整列宽和格式
            worksheet = writer.sheets['案例数据']
            
            # 设置列宽
            worksheet.column_dimensions['A'].width = 30  # 案例标题
            worksheet.column_dimensions['B'].width = 60  # 案例内容
            worksheet.column_dimensions['C'].width = 15  # 案例日期
            worksheet.column_dimensions['D'].width = 60  # 法官判决
            
            # 设置第一行为粗体（标题行）
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 添加说明
            from openpyxl import Workbook
            if '说明' not in writer.sheets:
                ws_instructions = writer.book.create_sheet('说明', 0)
                instructions = [
                    ['案例导入模板使用说明'],
                    [''],
                    ['列说明：'],
                    ['1. 案例标题（必填）：案例的名称或标题'],
                    ['2. 案例内容（必填）：案例的详细内容描述'],
                    ['3. 案例日期（可选）：案例发生的日期，格式：YYYY-MM-DD'],
                    ['4. 法官判决（可选）：法官的实际判决内容'],
                    [''],
                    ['注意事项：'],
                    ['- 第一行为标题行，请勿删除'],
                    ['- 案例标题和案例内容为必填项'],
                    ['- 可以删除示例数据，添加自己的案例'],
                    ['- 日期格式必须为 YYYY-MM-DD（如：2024-01-15）'],
                    ['- 支持批量导入多个案例'],
                ]
                
                for row_idx, row_data in enumerate(instructions, start=1):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        cell = ws_instructions.cell(row=row_idx, column=col_idx)
                        cell.value = cell_value
                        if row_idx == 1:  # 标题行
                            cell.font = Font(bold=True, size=14)
                        elif row_idx <= 2:
                            pass
                        elif '列说明' in str(cell_value) or '注意事项' in str(cell_value):
                            cell.font = Font(bold=True)
                
                ws_instructions.column_dimensions['A'].width = 80
        
        return filepath
    
    def export_questions(self, questions: List[str], questions_by_case: List[Dict] = None, filename: str = None) -> str:
        """
        导出问题列表到Excel
        
        Args:
            questions: 问题列表（简单格式）
            questions_by_case: 带案例信息的问题列表（可选），格式为 [{'case_id': '...', 'case_title': '...', 'question': '...'}]
            filename: 输出文件名，如果不提供则自动生成
            
        Returns:
            生成的Excel文件路径
        """
        if not questions and not questions_by_case:
            raise ValueError("问题列表为空")
        
        # 准备数据
        data = []
        
        if questions_by_case:
            # 使用带案例信息的数据
            for index, item in enumerate(questions_by_case, start=1):
                row = {
                    '序号': index,
                    '案例ID': item.get('case_id', ''),
                    '案例标题': item.get('case_title', ''),
                    '问题': item.get('question', '')
                }
                data.append(row)
        else:
            # 使用简单格式
            for index, question in enumerate(questions, start=1):
                row = {
                    '序号': index,
                    '问题': question
                }
                data.append(row)
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'generated_questions_{timestamp}.xlsx'
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.results_dir, filename)
        
        # 导出到Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='生成的问题', index=False)
            
            # 获取工作表对象以调整列宽
            worksheet = writer.sheets['生成的问题']
            
            # 自动调整列宽
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # 设置标题行样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置列宽
            if questions_by_case:
                worksheet.column_dimensions['A'].width = 10  # 序号
                worksheet.column_dimensions['B'].width = 15  # 案例ID
                worksheet.column_dimensions['C'].width = 30  # 案例标题
                worksheet.column_dimensions['D'].width = 80  # 问题
                
                # 设置问题列自动换行
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    row[3].alignment = Alignment(wrap_text=True, vertical='top')
            else:
                worksheet.column_dimensions['A'].width = 10  # 序号
                worksheet.column_dimensions['B'].width = 80  # 问题
                
                # 设置问题列自动换行
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    row[1].alignment = Alignment(wrap_text=True, vertical='top')
        
        return filepath
    
    def export_case_list(self, cases: List[Dict], filename: str = None, include_masked: bool = True) -> str:
        """
        导出案例列表到Excel（包含脱敏列）
        
        Args:
            cases: 案例列表
            filename: 输出文件名，如果不提供则自动生成
            include_masked: 是否包含脱敏列
            
        Returns:
            生成的Excel文件路径
        """
        if not cases:
            raise ValueError("案例列表为空")
        
        # 准备数据
        data = []
        for case in cases:
            judge_decision = case.get('judge_decision', '')
            if judge_decision == 'nan' or (isinstance(judge_decision, float) and pd.isna(judge_decision)):
                judge_decision = ''
            
            row = {
                '案例ID': case.get('id', ''),
                '案例标题': case.get('title', ''),
                '案例内容': case.get('case_text', ''),
                '法官判决': judge_decision,
                '案例日期': case.get('case_date', ''),
                '创建时间': case.get('created_at', '')
            }
            
            # 如果需要脱敏列，添加脱敏后的内容（使用API脱敏）
            if include_masked:
                masked_case = self.data_masker.mask_case_with_api(case)
                row['案例内容（脱敏）'] = masked_case.get('case_text_masked', '')
                row['法官判决（脱敏）'] = masked_case.get('judge_decision_masked', '')
            
            data.append(row)
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'案例列表_{timestamp}.xlsx'
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.results_dir, filename)
        
        # 导出到Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='案例列表', index=False)
            
            # 获取工作表对象以调整列宽
            worksheet = writer.sheets['案例列表']
            
            # 自动调整列宽
            from openpyxl.utils import get_column_letter
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                )
                # 设置最大宽度限制
                adjusted_width = min(max_length + 2, 100)
                column_letter = get_column_letter(idx + 1)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        return filepath


# 创建全局实例
excel_exporter = ExcelExporter()

