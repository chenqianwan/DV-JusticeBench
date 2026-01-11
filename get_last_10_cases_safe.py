#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
获取最后10个案例的名称 - 安全版本
"""
import sys
import os

def main():
    file_path = 'data/108个案例_新标准评估_完整版_最终版.xlsx'
    
    # 尝试不同的方法
    methods = [
        ('pandas + openpyxl', lambda: __import__('pandas').read_excel(file_path, engine='openpyxl')),
        ('pandas + xlrd', lambda: __import__('pandas').read_excel(file_path, engine='xlrd')),
        ('openpyxl直接读取', lambda: read_with_openpyxl(file_path)),
    ]
    
    for method_name, method_func in methods:
        try:
            print(f"尝试方法: {method_name}...")
            if method_name == 'openpyxl直接读取':
                df = method_func()
            else:
                df = method_func()
            
            if df is not None and len(df) > 0:
                print(f"✓ 成功读取，总行数: {len(df)}")
                
                if '案例ID' in df.columns and '案例标题' in df.columns:
                    case_info = df[['案例ID', '案例标题']].drop_duplicates()
                    case_info = case_info.sort_values('案例ID')
                    last_10 = case_info.tail(10)
                    
                    print("\n" + "=" * 80)
                    print("最后10个案例:")
                    print("=" * 80)
                    
                    for idx, (_, row) in enumerate(last_10.iterrows(), 1):
                        print(f"{idx:2d}. {row['案例ID']}")
                        print(f"    标题: {row['案例标题']}")
                        print()
                    return
                else:
                    print(f"列名: {list(df.columns[:10])}")
            break
        except Exception as e:
            print(f"  ✗ 失败: {e}")
            continue
    
    print("\n所有方法都失败了")

def read_with_openpyxl(file_path):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        
        # 读取第一行作为列名
        headers = [cell.value for cell in ws[1]]
        
        # 找到案例ID和案例标题的列索引
        try:
            case_id_idx = headers.index('案例ID')
            case_title_idx = headers.index('案例标题')
        except ValueError:
            return None
        
        # 读取所有数据
        cases = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[case_id_idx] and row[case_title_idx]:
                case_id = row[case_id_idx]
                case_title = row[case_title_idx]
                if case_id not in cases:
                    cases[case_id] = case_title
        
        # 转换为DataFrame格式
        import pandas as pd
        df = pd.DataFrame(list(cases.items()), columns=['案例ID', '案例标题'])
        return df
        
    except Exception as e:
        print(f"openpyxl读取失败: {e}")
        return None

if __name__ == '__main__':
    main()
