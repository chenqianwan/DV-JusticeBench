"""
Flask主应用
提供Web界面和API接口
"""
from flask import Flask, render_template, request, jsonify, send_file
import os
import json
import uuid
import threading
import concurrent.futures
from datetime import datetime
from threading import Lock
from utils.ai_api import ai_api, UnifiedAIAPI
from utils.case_manager import case_manager
from utils.excel_export import excel_exporter
from utils.excel_import import ExcelImporter
from utils.similarity import similarity_calculator
from utils.process_cleanup import setup_signal_handlers, SafeThreadPoolExecutor
from utils.doc_reader import DocReader
from utils.data_masking import DataMasker, DataMaskerAPI
from utils.evaluator import AnswerEvaluator
from config import RESULTS_DIR, MAX_CONCURRENT_WORKERS
from werkzeug.utils import secure_filename
import tempfile

# 设置信号处理器，确保中断时正确清理
setup_signal_handlers()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key')

# 存储分析结果（实际应用中应使用数据库）
analysis_results = []

# Excel导入器
excel_importer = ExcelImporter()

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# 全局进度跟踪字典
batch_progress = {}
progress_lock = Lock()

# V2版本会话管理（内存字典）
sessions_v2 = {}
sessions_lock = Lock()

# 初始化工具
doc_reader = DocReader()
data_masker = DataMasker()
data_masker_api = DataMaskerAPI()


@app.route('/')
def index():
    """主页面"""
    return render_template('index.html')


@app.route('/api/cases', methods=['GET'])
def get_cases():
    """获取所有案例"""
    try:
        cases = case_manager.get_all_cases()
        return jsonify({'success': True, 'cases': cases})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/cases', methods=['POST'])
def add_case():
    """添加新案例"""
    try:
        data = request.json
        case_id = case_manager.add_case(
            title=data.get('title', '未命名案例'),
            case_text=data.get('case_text', ''),
            judge_decision=data.get('judge_decision', ''),
            case_date=data.get('case_date', ''),
            metadata=data.get('metadata', {})
        )
        return jsonify({'success': True, 'case_id': case_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/cases/<case_id>', methods=['GET'])
def get_case(case_id):
    """获取指定案例"""
    try:
        case = case_manager.get_case(case_id)
        if case:
            return jsonify({'success': True, 'case': case})
        else:
            return jsonify({'success': False, 'error': '案例不存在'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/cases/<case_id>', methods=['DELETE'])
def delete_case(case_id):
    """删除案例"""
    try:
        success = case_manager.delete_case(case_id)
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': '案例不存在'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/generate_questions', methods=['POST'])
def generate_questions():
    """生成测试问题（单个案例）"""
    try:
        data = request.json
        case_id = data.get('case_id')
        num_questions = data.get('num_questions', 10)
        
        if not case_id:
            return jsonify({'success': False, 'error': '缺少案例ID'}), 400
        
        case = case_manager.get_case(case_id)
        if not case:
            return jsonify({'success': False, 'error': '案例不存在'}), 404
        
        case_text = case.get('case_text', '')
        if not case_text:
            return jsonify({'success': False, 'error': '案例内容为空'}), 400
        
        # 调用DeepSeek API生成问题
        questions = ai_api.generate_questions(case_text, num_questions)
        
        return jsonify({
            'success': True,
            'questions': questions,
            'case_id': case_id
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/generate_questions_batch', methods=['POST'])
def generate_questions_batch():
    """批量生成问题（多个案例，每个案例生成指定数量的问题）"""
    try:
        data = request.json
        case_ids = data.get('case_ids', [])
        num_questions_per_case = data.get('num_questions_per_case', 10)
        
        if not case_ids:
            return jsonify({'success': False, 'error': '缺少案例ID列表'}), 400
        
        if len(case_ids) == 0:
            return jsonify({'success': False, 'error': '案例ID列表为空'}), 400
        
        # 获取所有案例
        cases = []
        for case_id in case_ids:
            case = case_manager.get_case(case_id)
            if case:
                cases.append((case_id, case))
        
        if not cases:
            return jsonify({'success': False, 'error': '未找到有效的案例'}), 404
        
        # 为每个案例生成问题
        all_questions = []
        questions_by_case = []
        
        for case_id, case_data in cases:
            case_text = case_data.get('case_text', '')
            if not case_text:
                continue
            
            try:
                # 为每个案例生成指定数量的问题
                questions = deepseek_api.generate_questions(case_text, num_questions_per_case)
                
                # 为每个问题添加案例信息
                for q in questions:
                    all_questions.append(q)
                    questions_by_case.append({
                        'case_id': case_id,
                        'case_title': case_data.get('title', ''),
                        'question': q
                    })
            except Exception as e:
                print(f"为案例 {case_data.get('title', case_id)} 生成问题失败: {str(e)}")
                continue
        
        return jsonify({
            'success': True,
            'all_questions': all_questions,
            'questions_by_case': questions_by_case,
            'total_questions': len(all_questions),
            'total_cases': len(cases)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/analyze', methods=['POST'])
def analyze():
    """AI分析案例"""
    try:
        data = request.json
        case_id = data.get('case_id')
        question = data.get('question', '')
        
        if not case_id:
            return jsonify({'success': False, 'error': '缺少案例ID'}), 400
        
        case = case_manager.get_case(case_id)
        if not case:
            return jsonify({'success': False, 'error': '案例不存在'}), 404
        
        case_text = case.get('case_text', '')
        if not case_text:
            return jsonify({'success': False, 'error': '案例内容为空'}), 400
        
        # 调用DeepSeek API分析
        ai_decision = ai_api.analyze_case(case_text, question)
        
        # 获取法官判决
        judge_decision = case.get('judge_decision', '')
        
        # 如果有法官判决，进行对比分析和相似度计算
        comparison = ''
        similarity_metrics = {}
        if judge_decision:
            try:
                comparison = ai_api.compare_decisions(ai_decision, judge_decision)
            except Exception as e:
                comparison = f'对比分析失败: {str(e)}'
            
            # 计算相似度指标
            try:
                similarity_metrics = similarity_calculator.calculate_metrics(
                    ai_decision, judge_decision
                )
            except Exception as e:
                print(f"相似度计算失败: {str(e)}")
                similarity_metrics = {
                    'overall_similarity': 0.0,
                    'keyword_similarity': 0.0,
                    'result_consistency': 0.0,
                    'legal_basis_similarity': 0.0,
                    'reasoning_similarity': 0.0,
                    'has_judge_decision': True
                }
        else:
            similarity_metrics = {
                'overall_similarity': 0.0,
                'keyword_similarity': 0.0,
                'result_consistency': 0.0,
                'legal_basis_similarity': 0.0,
                'reasoning_similarity': 0.0,
                'has_judge_decision': False
            }
        
        # 保存分析结果
        result = {
            'case_id': case_id,
            'case_title': case.get('title', ''),
            'case_text': case_text,
            'question': question,
            'ai_decision': ai_decision,
            'judge_decision': judge_decision,
            'comparison': comparison,
            'similarity_metrics': similarity_metrics,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        analysis_results.append(result)
        
        return jsonify({
            'success': True,
            'result': result
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/results', methods=['GET'])
def get_results():
    """获取所有分析结果"""
    try:
        return jsonify({'success': True, 'results': analysis_results})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/analyze_batch', methods=['POST'])
def analyze_batch():
    """批量分析多个案例（使用线程池并发处理，固定50并发）"""
    try:
        data = request.json
        case_ids = data.get('case_ids', [])
        question = data.get('question', '')
        
        if not case_ids:
            return jsonify({'success': False, 'error': '缺少案例ID列表'}), 400
        
        if len(case_ids) == 0:
            return jsonify({'success': False, 'error': '案例ID列表为空'}), 400
        
        # 生成任务ID
        task_id = str(uuid.uuid4())
        
        # 获取所有案例
        cases = []
        for case_id in case_ids:
            case = case_manager.get_case(case_id)
            if case:
                cases.append((case_id, case))
        
        if not cases:
            return jsonify({'success': False, 'error': '未找到有效的案例'}), 404
        
        # 初始化进度
        with progress_lock:
            batch_progress[task_id] = {
                'total': len(cases),
                'completed': 0,
                'success': 0,
                'failed': 0,
                'status': 'running',  # running, completed, failed
                'results': [],
                'errors': [],
                'start_time': datetime.now().isoformat()
            }
        
        # 使用配置的并发数（固定50）
        max_workers = min(MAX_CONCURRENT_WORKERS, len(cases))  # 不超过案例数量
        
        print(f"[批量分析] 任务 {task_id} 开始处理 {len(cases)} 个案例，并发数: {max_workers}")
        
        # 定义单个案例的分析函数
        def analyze_single_case(case_id, case_data):
            """分析单个案例的辅助函数"""
            try:
                case_text = case_data.get('case_text', '')
                if not case_text:
                    return {'success': False, 'case_id': case_id, 'error': '案例内容为空'}
                
                # 调用DeepSeek API分析
                ai_decision = ai_api.analyze_case(case_text, question)
                
                # 获取法官判决
                judge_decision = case_data.get('judge_decision', '')
                
                # 如果有法官判决，进行对比分析和相似度计算
                comparison = ''
                similarity_metrics = {}
                if judge_decision:
                    try:
                        comparison = ai_api.compare_decisions(ai_decision, judge_decision)
                    except Exception as e:
                        comparison = f'对比分析失败: {str(e)}'
                    
                    # 计算相似度指标
                    try:
                        similarity_metrics = similarity_calculator.calculate_metrics(
                            ai_decision, judge_decision
                        )
                    except Exception as e:
                        print(f"相似度计算失败: {str(e)}")
                        similarity_metrics = {
                            'overall_similarity': 0.0,
                            'keyword_similarity': 0.0,
                            'result_consistency': 0.0,
                            'legal_basis_similarity': 0.0,
                            'reasoning_similarity': 0.0,
                            'has_judge_decision': True
                        }
                else:
                    similarity_metrics = {
                        'overall_similarity': 0.0,
                        'keyword_similarity': 0.0,
                        'result_consistency': 0.0,
                        'legal_basis_similarity': 0.0,
                        'reasoning_similarity': 0.0,
                        'has_judge_decision': False
                    }
                
                # 构建结果
                result = {
                    'case_id': case_id,
                    'case_title': case_data.get('title', ''),
                    'case_text': case_text,
                    'question': question,
                    'ai_decision': ai_decision,
                    'judge_decision': judge_decision,
                    'comparison': comparison,
                    'similarity_metrics': similarity_metrics,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                # 保存分析结果
                analysis_results.append(result)
                
                # 更新进度
                with progress_lock:
                    if task_id in batch_progress:
                        batch_progress[task_id]['completed'] += 1
                        batch_progress[task_id]['success'] += 1
                        batch_progress[task_id]['results'].append(result)
                
                return {'success': True, 'result': result}
                
            except Exception as e:
                # 更新失败计数
                with progress_lock:
                    if task_id in batch_progress:
                        batch_progress[task_id]['completed'] += 1
                        batch_progress[task_id]['failed'] += 1
                        batch_progress[task_id]['errors'].append({
                            'case_id': case_id,
                            'case_title': case_data.get('title', ''),
                            'error': str(e)
                        })
                
                return {'success': False, 'case_id': case_id, 'error': str(e)}
        
        # 使用线程池并发处理
        def run_analysis():
            """在后台线程中运行分析"""
            try:
                with SafeThreadPoolExecutor(max_workers=max_workers) as executor:
                    # 提交所有任务
                    future_to_case = {
                        executor.submit(analyze_single_case, case_id, case_data): (case_id, case_data)
                        for case_id, case_data in cases
                    }
                    
                    # 收集结果
                    for future in concurrent.futures.as_completed(future_to_case):
                        case_id, case_data = future_to_case[future]
                        try:
                            result = future.result()
                            # 进度已在 analyze_single_case 中更新
                        except Exception as e:
                            print(f"[批量分析] 异常: {case_data.get('title', case_id)} - {str(e)}")
                
                # 标记完成
                with progress_lock:
                    if task_id in batch_progress:
                        batch_progress[task_id]['status'] = 'completed'
                        batch_progress[task_id]['end_time'] = datetime.now().isoformat()
                        
                        # 按原始顺序排序结果
                        case_id_order = {case_id: idx for idx, (case_id, _) in enumerate(cases)}
                        batch_progress[task_id]['results'].sort(
                            key=lambda x: case_id_order.get(x['case_id'], 999)
                        )
                
                print(f"[批量分析] 任务 {task_id} 完成")
            except Exception as e:
                with progress_lock:
                    if task_id in batch_progress:
                        batch_progress[task_id]['status'] = 'failed'
                        batch_progress[task_id]['error'] = str(e)
                print(f"[批量分析] 任务 {task_id} 失败: {str(e)}")
        
        # 在后台线程中启动分析
        analysis_thread = threading.Thread(target=run_analysis, daemon=True)
        analysis_thread.start()
        
        # 立即返回任务ID
        return jsonify({
            'success': True,
            'task_id': task_id,
            'message': '批量分析已启动'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/batch_progress/<task_id>', methods=['GET'])
def get_batch_progress(task_id):
    """获取批量分析进度"""
    try:
        with progress_lock:
            if task_id not in batch_progress:
                return jsonify({
                    'success': False,
                    'error': '任务不存在或已过期'
                }), 404
            
            progress = batch_progress[task_id].copy()
            
            # 计算百分比
            if progress['total'] > 0:
                progress['percentage'] = int((progress['completed'] / progress['total']) * 100)
            else:
                progress['percentage'] = 0
            
            # 计算预计剩余时间
            if progress['status'] == 'running' and progress['completed'] > 0:
                elapsed = (datetime.now() - datetime.fromisoformat(progress['start_time'])).total_seconds()
                if elapsed > 0:
                    avg_time_per_case = elapsed / progress['completed']
                    remaining_cases = progress['total'] - progress['completed']
                    estimated_remaining = avg_time_per_case * remaining_cases
                    progress['estimated_remaining_seconds'] = int(estimated_remaining)
                else:
                    progress['estimated_remaining_seconds'] = None
            else:
                progress['estimated_remaining_seconds'] = None
        
        return jsonify({
            'success': True,
            'progress': progress
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/export_excel', methods=['POST'])
def export_excel():
    """导出分析结果到Excel"""
    try:
        data = request.json
        results = data.get('results', analysis_results)
        
        if not results:
            return jsonify({'success': False, 'error': '没有可导出的结果'}), 400
        
        # 生成Excel文件
        filepath = excel_exporter.export_analysis_results(results)
        
        # 返回文件
        return send_file(
            filepath,
            as_attachment=True,
            download_name=os.path.basename(filepath),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/export_cases', methods=['POST'])
def export_cases():
    """导出案例摘要到Excel"""
    try:
        cases = case_manager.get_all_cases()
        
        if not cases:
            return jsonify({'success': False, 'error': '没有可导出的案例'}), 400
        
        # 生成Excel文件
        filepath = excel_exporter.export_case_summary(cases)
        
        # 返回文件
        return send_file(
            filepath,
            as_attachment=True,
            download_name=os.path.basename(filepath),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/api/download_template', methods=['GET'])
def download_template():
    """下载Excel导入模板"""
    try:
        # 生成模板文件
        filepath = excel_exporter.generate_import_template()
        
        # 返回文件
        return send_file(
            filepath,
            as_attachment=True,
            download_name='案例导入模板.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/export_questions', methods=['POST'])
def export_questions():
    """导出生成的问题到Excel"""
    try:
        data = request.json
        questions = data.get('questions', [])
        questions_by_case = data.get('questions_by_case', None)
        
        if not questions and not questions_by_case:
            return jsonify({'success': False, 'error': '没有可导出的问题'}), 400
        
        # 生成Excel文件
        filepath = excel_exporter.export_questions(questions, questions_by_case)
        
        # 返回文件
        return send_file(
            filepath,
            as_attachment=True,
            download_name=os.path.basename(filepath),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/import_cases', methods=['POST'])
def import_cases():
    """批量导入案例（从Excel文件）"""
    try:
        # 检查文件是否存在
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有上传文件'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': '未选择文件'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': '不支持的文件格式，请上传 .xlsx 或 .xls 文件'}), 400
        
        # 保存临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            tmp_filepath = tmp_file.name
        
        try:
            # 解析Excel文件
            cases = excel_importer.parse_excel(tmp_filepath)
            
            if not cases:
                return jsonify({'success': False, 'error': 'Excel文件中没有找到有效的案例数据'}), 400
            
            # 验证数据
            validation = excel_importer.validate_cases(cases)
            valid_cases = validation['valid']
            invalid_cases = validation['invalid']
            
            if not valid_cases:
                return jsonify({
                    'success': False,
                    'error': '没有有效的案例数据',
                    'invalid_cases': invalid_cases
                }), 400
            
            # 批量导入有效案例
            imported_count = 0
            imported_cases = []
            for case in valid_cases:
                try:
                    case_id = case_manager.add_case(
                        title=case['title'],
                        case_text=case['case_text'],
                        judge_decision=case.get('judge_decision', ''),
                        case_date=case.get('case_date', ''),
                        metadata={}
                    )
                    imported_cases.append({
                        'case_id': case_id,
                        'title': case['title']
                    })
                    imported_count += 1
                except Exception as e:
                    print(f"导入案例失败: {str(e)}")
            
            return jsonify({
                'success': True,
                'imported_count': imported_count,
                'total_count': len(cases),
                'invalid_count': len(invalid_cases),
                'invalid_cases': invalid_cases,
                'imported_cases': imported_cases
            })
            
        finally:
            # 删除临时文件
            try:
                os.unlink(tmp_filepath)
            except:
                pass
                
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== V2 API端点（新版4步骤流程） ====================

@app.route('/v2')
def index_v2():
    """V2版本主页面"""
    return render_template('index_v2.html')


@app.route('/api/v2/upload', methods=['POST'])
def upload_v2():
    """上传.docx文件并提取文本"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '未找到上传文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '未选择文件'}), 400
        
        # 检查文件扩展名
        if not file.filename.lower().endswith('.docx'):
            return jsonify({'success': False, 'error': '只支持.docx文件'}), 400
        
        # 保存临时文件
        temp_dir = tempfile.gettempdir()
        filename = secure_filename(file.filename)
        filepath = os.path.join(temp_dir, f"{uuid.uuid4()}_{filename}")
        file.save(filepath)
        
        try:
            # 提取文本
            text = doc_reader.read_doc(filepath)
            
            if not text or len(text.strip()) == 0:
                return jsonify({'success': False, 'error': '文件内容为空'}), 400
            
            # 创建新会话
            session_id = str(uuid.uuid4())
            with sessions_lock:
                sessions_v2[session_id] = {
                    'step': 1,
                    'original_text': text,
                    'masked_text': '',
                    'questions': [],
                    'model': 'deepseek',
                    'answers': [],
                    'evaluations': [],
                    'created_at': datetime.now().isoformat()
                }
            
            return jsonify({
                'success': True,
                'session_id': session_id,
                'text': text,
                'text_length': len(text)
            })
            
        finally:
            # 删除临时文件
            try:
                os.unlink(filepath)
            except:
                pass
                
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/v2/mask', methods=['POST'])
def mask_v2():
    """隐私脱敏"""
    try:
        data = request.json
        session_id = data.get('session_id')
        mode = data.get('mode', 'fast')  # 'fast' 或 'review'
        text = data.get('text', '')
        
        if not session_id:
            return jsonify({'success': False, 'error': '缺少session_id'}), 400
        
        with sessions_lock:
            if session_id not in sessions_v2:
                return jsonify({'success': False, 'error': '会话不存在'}), 404
            
            session = sessions_v2[session_id]
            
            # 如果提供了text（审核模式下用户编辑后的文本），使用它
            if text:
                original_text = text
            else:
                original_text = session['original_text']
        
        # 根据模式选择脱敏方式
        if mode == 'fast':
            # 快速模式：使用正则表达式脱敏
            masked_text = data_masker.mask_text(original_text)
        else:
            # 审核模式：使用API脱敏（更准确）
            masked_text = data_masker_api.mask_text_with_api(original_text)
            if not masked_text:
                # API失败时回退到正则脱敏
                masked_text = data_masker.mask_text(original_text)
        
        # 更新会话
        with sessions_lock:
            sessions_v2[session_id]['masked_text'] = masked_text
            sessions_v2[session_id]['step'] = 2
        
        return jsonify({
            'success': True,
            'masked_text': masked_text,
            'mode': mode
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/v2/generate_questions', methods=['POST'])
def generate_questions_v2():
    """生成问题（固定5个）"""
    try:
        data = request.json
        session_id = data.get('session_id')
        
        if not session_id:
            return jsonify({'success': False, 'error': '缺少session_id'}), 400
        
        with sessions_lock:
            if session_id not in sessions_v2:
                return jsonify({'success': False, 'error': '会话不存在'}), 404
            
            session = sessions_v2[session_id]
            masked_text = session['masked_text']
        
        if not masked_text:
            return jsonify({'success': False, 'error': '缺少脱敏文本'}), 400
        
        # 使用DeepSeek生成5个问题
        api = UnifiedAIAPI(provider='deepseek')
        questions = api.generate_questions(masked_text, num_questions=5)
        
        # 更新会话
        with sessions_lock:
            sessions_v2[session_id]['questions'] = questions
            sessions_v2[session_id]['step'] = 3
        
        return jsonify({
            'success': True,
            'questions': questions
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/v2/update_questions', methods=['POST'])
def update_questions_v2():
    """更新问题列表（支持编辑/删除/添加）"""
    try:
        data = request.json
        session_id = data.get('session_id')
        questions = data.get('questions', [])
        
        if not session_id:
            return jsonify({'success': False, 'error': '缺少session_id'}), 400
        
        with sessions_lock:
            if session_id not in sessions_v2:
                return jsonify({'success': False, 'error': '会话不存在'}), 404
            
            sessions_v2[session_id]['questions'] = questions
        
        return jsonify({'success': True})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/v2/generate_answer', methods=['POST'])
def generate_answer_v2():
    """生成单个问题的AI答案"""
    try:
        data = request.json
        session_id = data.get('session_id')
        question_index = data.get('question_index')
        model = data.get('model', 'deepseek')
        use_thinking = data.get('use_thinking', True)
        
        if not session_id:
            return jsonify({'success': False, 'error': '缺少session_id'}), 400
        
        if question_index is None:
            return jsonify({'success': False, 'error': '缺少question_index'}), 400
        
        with sessions_lock:
            if session_id not in sessions_v2:
                return jsonify({'success': False, 'error': '会话不存在'}), 404
            
            session = sessions_v2[session_id]
            masked_text = session['masked_text']
            questions = session['questions']
            
            if question_index >= len(questions):
                return jsonify({'success': False, 'error': '问题索引超出范围'}), 400
            
            question = questions[question_index]
        
        # 映射前端模型名称到实际provider和model
        provider_map = {
            'deepseek': ('deepseek', None),
            'deepseek-thinking': ('deepseek', None),
            'gpt': ('chatgpt', 'gpt-4o'),
            'claude': ('claude', 'claude-opus-4-20250514'),
            'gemini': ('chatgpt', 'gemini-1.5-pro-002'),
            'qwen': ('qwen', 'qwen-max')
        }
        
        provider, model_name = provider_map.get(model, ('deepseek', None))
        
        # 生成答案
        api = UnifiedAIAPI(provider=provider, model=model_name)
        result = api.analyze_case(masked_text, question, use_thinking=use_thinking)
        
        answer = result.get('answer', '')
        reasoning = result.get('reasoning', '')
        
        # 确保answers列表足够长
        with sessions_lock:
            while len(sessions_v2[session_id]['answers']) <= question_index:
                sessions_v2[session_id]['answers'].append({})
            
            sessions_v2[session_id]['answers'][question_index] = {
                'question': question,
                'answer': answer,
                'reasoning': reasoning,
                'model': model
            }
            sessions_v2[session_id]['model'] = model
        
        return jsonify({
            'success': True,
            'answer': answer,
            'reasoning': reasoning
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/v2/evaluate', methods=['POST'])
def evaluate_v2():
    """评分单个答案（5维度+错误分类）"""
    try:
        data = request.json
        session_id = data.get('session_id')
        question_index = data.get('question_index')
        judge_decision = data.get('judge_decision', '')  # 法官判决（可选）
        
        if not session_id:
            return jsonify({'success': False, 'error': '缺少session_id'}), 400
        
        if question_index is None:
            return jsonify({'success': False, 'error': '缺少question_index'}), 400
        
        with sessions_lock:
            if session_id not in sessions_v2:
                return jsonify({'success': False, 'error': '会话不存在'}), 404
            
            session = sessions_v2[session_id]
            masked_text = session['masked_text']
            answers = session['answers']
            
            if question_index >= len(answers):
                return jsonify({'success': False, 'error': '答案索引超出范围'}), 400
            
            answer_data = answers[question_index]
            question = answer_data['question']
            answer = answer_data['answer']
        
        # 评分
        evaluator = AnswerEvaluator()
        evaluation_result = evaluator.evaluate_answer(
            ai_answer=answer,
            judge_decision=judge_decision,
            question=question,
            case_text=masked_text
        )
        
        # 调试日志
        print(f"\n=== 评分结果调试 ===")
        print(f"评分结果键: {evaluation_result.keys()}")
        print(f"总分: {evaluation_result.get('总分', 'NOT FOUND')}")
        print(f"各维度得分: {evaluation_result.get('各维度得分', 'NOT FOUND')}")
        print("==================\n")
        
        # 转换为前端期望的格式（英文键名）
        evaluation = {
            'total_score': evaluation_result.get('总分', 0),
            'percentage': evaluation_result.get('百分制', 0),
            'scores': evaluation_result.get('各维度得分', {}),
            'detailed_evaluation': evaluation_result.get('详细评价', ''),
            'error_flags': evaluation_result.get('错误标记', ''),
            'error_counts': {},  # 需要从错误详情中提取
            'grade': evaluation_result.get('分档', ''),
            'thinking': evaluation_result.get('评价Thinking', '')
        }
        
        # 提取错误计数
        error_details = evaluation_result.get('错误详情', {})
        evaluation['error_counts'] = {
            'major_errors': len(error_details.get('重大错误', [])),
            'obvious_errors': len(error_details.get('明显错误', [])),
            'minor_errors': len(error_details.get('微小错误', [])),
            'abandoned_law_citations': 0  # 如果有的话
        }
        
        print(f"转换后的evaluation: {evaluation}")
        print(f"scores字段: {evaluation['scores']}")
        
        # 确保evaluations列表足够长
        with sessions_lock:
            while len(sessions_v2[session_id]['evaluations']) <= question_index:
                sessions_v2[session_id]['evaluations'].append({})
            
            sessions_v2[session_id]['evaluations'][question_index] = evaluation
            sessions_v2[session_id]['step'] = 4
        
        return jsonify({
            'success': True,
            'evaluation': evaluation
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/v2/session/<session_id>', methods=['GET'])
def get_session_v2(session_id):
    """获取会话状态"""
    try:
        with sessions_lock:
            if session_id not in sessions_v2:
                return jsonify({'success': False, 'error': '会话不存在'}), 404
            
            session = sessions_v2[session_id].copy()
        
        return jsonify({
            'success': True,
            'session': session
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/v2/export_excel', methods=['POST'])
def export_excel_v2():
    """导出评估结果到Excel"""
    try:
        data = request.json
        session_id = data.get('session_id')
        
        if not session_id:
            return jsonify({'success': False, 'error': '缺少session_id'}), 400
        
        with sessions_lock:
            if session_id not in sessions_v2:
                return jsonify({'success': False, 'error': '会话不存在'}), 404
            
            session = sessions_v2[session_id]
        
        # 创建Excel
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        # 准备数据
        excel_data = []
        for i, (question, answer, evaluation) in enumerate(zip(
            session['questions'],
            session['answers'],
            session['evaluations']
        )):
            scores = evaluation.get('scores', {})
            errors = evaluation.get('errors', {})
            
            row = {
                'Question No.': i + 1,
                'Question': question,
                'AI Answer': answer['answer'],
                'Total Score': evaluation.get('total_score', 0),
                'Legal Fact Coverage': scores.get('legal_fact_coverage', 0),
                'Subsumption Chain': scores.get('subsumption_chain', 0),
                'Value & Empathy': scores.get('value_empathy', 0),
                'Key Facts Coverage': scores.get('key_facts_coverage', 0),
                'Outcome Consistency': scores.get('outcome_consistency', 0),
                'Major Errors': errors.get('major', 0),
                'Obvious Errors': errors.get('obvious', 0),
                'Minor Errors': errors.get('minor', 0),
                'Abandoned Law Citations': errors.get('abandoned_law_citations', 0)
            }
            excel_data.append(row)
        
        # 创建DataFrame
        df = pd.DataFrame(excel_data)
        
        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Evaluation Results', index=False)
            
            # 美化Excel
            worksheet = writer.sheets['Evaluation Results']
            
            # 设置标题行样式
            header_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 设置列宽
            column_widths = {
                'A': 12,  # Question No.
                'B': 60,  # Question
                'C': 80,  # AI Answer
                'D': 12,  # Total Score
                'E': 18,  # Legal Fact Coverage
                'F': 18,  # Subsumption Chain
                'G': 18,  # Value & Empathy
                'H': 18,  # Key Facts Coverage
                'I': 18,  # Outcome Consistency
                'J': 12,  # Major Errors
                'K': 15,  # Obvious Errors
                'L': 12,  # Minor Errors
                'M': 20   # Abandoned Law Citations
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 设置单元格对齐和换行
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                row[1].alignment = Alignment(wrap_text=True, vertical='top')  # Question
                row[2].alignment = Alignment(wrap_text=True, vertical='top')  # AI Answer
                for cell in row[3:]:  # Scores
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 添加摘要工作表
            summary_data = {
                'Metric': [
                    'Model',
                    'Total Questions',
                    'Average Total Score',
                    'Average Legal Fact Coverage',
                    'Average Subsumption Chain',
                    'Average Value & Empathy',
                    'Average Key Facts Coverage',
                    'Average Outcome Consistency',
                    'Total Major Errors',
                    'Total Obvious Errors',
                    'Total Minor Errors',
                    'Total Abandoned Law Citations'
                ],
                'Value': [
                    session.get('model', 'N/A'),
                    len(session['questions']),
                    round(sum(e.get('total_score', 0) for e in session['evaluations']) / len(session['evaluations']), 2) if session['evaluations'] else 0,
                    round(sum(e.get('scores', {}).get('legal_fact_coverage', 0) for e in session['evaluations']) / len(session['evaluations']), 2) if session['evaluations'] else 0,
                    round(sum(e.get('scores', {}).get('subsumption_chain', 0) for e in session['evaluations']) / len(session['evaluations']), 2) if session['evaluations'] else 0,
                    round(sum(e.get('scores', {}).get('value_empathy', 0) for e in session['evaluations']) / len(session['evaluations']), 2) if session['evaluations'] else 0,
                    round(sum(e.get('scores', {}).get('key_facts_coverage', 0) for e in session['evaluations']) / len(session['evaluations']), 2) if session['evaluations'] else 0,
                    round(sum(e.get('scores', {}).get('outcome_consistency', 0) for e in session['evaluations']) / len(session['evaluations']), 2) if session['evaluations'] else 0,
                    sum(e.get('errors', {}).get('major', 0) for e in session['evaluations']),
                    sum(e.get('errors', {}).get('obvious', 0) for e in session['evaluations']),
                    sum(e.get('errors', {}).get('minor', 0) for e in session['evaluations']),
                    sum(e.get('errors', {}).get('abandoned_law_citations', 0) for e in session['evaluations'])
                ]
            }
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # 美化摘要工作表
            ws_summary = writer.sheets['Summary']
            for cell in ws_summary[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws_summary.column_dimensions['A'].width = 35
            ws_summary.column_dimensions['B'].width = 20
            
            for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row):
                row[1].alignment = Alignment(horizontal='center', vertical='center')
        
        output.seek(0)
        
        # 生成文件名
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        model_name = session.get('model', 'unknown').replace('-', '_')
        filename = f'{model_name}_evaluation_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)

